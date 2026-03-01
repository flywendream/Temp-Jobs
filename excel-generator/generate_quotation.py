#!/usr/bin/env python3
"""
服务报价单 Excel 生成器
根据模板结构生成带公式的服务报价单
"""

# 支持从项目 lib 目录加载依赖（用于 pip install --target ./lib）
import sys
from pathlib import Path
lib_path = Path(__file__).parent / 'lib'
if lib_path.exists():
    sys.path.insert(0, str(lib_path))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_quotation_template():
    """创建服务报价单 Excel 模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "服务报价单"

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题：服务报价单 (F2:I2 合并，居中对齐)
    ws.merge_cells('F2:I2')
    ws['F2'] = '服务报价单'
    ws['F2'].font = Font(bold=True, size=14)
    ws['F2'].alignment = Alignment(horizontal='center', vertical='center')

    # 左侧客户信息区域 (B5:E19)
    customer_labels = [
        ('B5', 'C5', '客户编号:'),
        ('B6', 'C6', '客户公司名称:'),
        ('B7', 'C7', 'Address:'),
        ('B8', 'C8', '客户公司地址:'),
        ('B9', 'C9', '客户联系人:'),
        ('B10', 'C10', '电话:'),
        ('B11', 'C11', '电子邮箱:'),
        ('B12', 'C12', '最终用户编号:'),
        ('B13', 'C13', '最终用户:'),
        ('B14', 'C14', '发货地址:'),
    ]
    for label_cell, _, label in customer_labels:
        ws[label_cell] = label
        ws[label_cell].alignment = Alignment(horizontal='left', vertical='center')
    # 合并输入区域 C5:E19 作为黄色备注区
    ws.merge_cells('C5:E19')
    ws['C5'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    ws['C5'].border = thin_border

    # 右侧报价详细信息 (F5:J20)
    quote_info = [
        ('F5', 'G5', '报价单号:'),
        ('F6', 'G6', '报价日期:'),
        ('F7', 'G7', '报价有效期:'),
        ('F8', 'G8', '税率:'),
        ('F9', 'G9', '类型:'),
        ('F10', 'G10', '账期:'),
    ]
    for label_cell, value_cell, label in quote_info:
        ws[label_cell] = label
        ws[label_cell].alignment = Alignment(horizontal='left', vertical='center')
    ws['H6'] = '2026-02-27'  # 报价日期
    ws['H7'] = '90天'
    ws['H8'] = '6%'
    ws['H9'] = 'Repair'
    ws['H10'] = '无'

    # 付款方式、交货期
    ws['F11'] = '付款方式:'
    ws.merge_cells('G11:J11')
    ws['G11'] = '100%发货前付清'
    ws.merge_cells('G12:J12')
    ws['G12'] = '收到全部款项且完成服务2周内,开具全额发票给买方'

    ws['F13'] = '交货期:'
    ws.merge_cells('G13:J13')
    ws['G13'] = '收到订单后,款到后约8周'

    # 联系信息
    ws['F14'] = '联系人:'
    ws['H14'] = '电话:'
    ws['F15'] = '电子邮箱:'
    ws['F16'] = '最终用户联系人:'
    ws['H16'] = '联系电话:'
    ws['F17'] = '电子邮箱:'

    # 设置右侧区域边框
    for row in range(5, 21):
        for col in range(6, 11):  # F-J
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # 明细项目表头 (B21:K21)
    headers = ['项目', '设备型号', '物料号', '内容描述', '说明', '数量', '未税优惠单价 (RMB)', '含税单价 (RMB)', '含税优惠单价 (RMB)', '含税优惠合计 (RMB)']
    for col, header in enumerate(headers, start=2):  # B=2
        cell = ws.cell(row=21, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 示例数据行 + 公式
    # 服务费用
    ws['B22'] = '服务费用'
    ws['B22'].font = Font(bold=True)
    ws.merge_cells('B22:K22')
    ws['B22'].border = thin_border

    detail_rows = [
        (23, '1', '', '', '', '', '', '', '', ''),
        (24, '2', '', '', '', '', '', '', '', ''),
    ]
    for row_num, *vals in detail_rows:
        ws.cell(row=row_num, column=2, value=vals[0])  # 项目编号
        ws.cell(row=row_num, column=3, value=vals[1])  # 设备型号
        ws.cell(row=row_num, column=4, value=vals[2])  # 物料号
        ws.cell(row=row_num, column=5, value=vals[3])  # 内容描述
        ws.cell(row=row_num, column=6, value=vals[4])  # 说明
        ws.cell(row=row_num, column=7, value=vals[5] if vals[5] else 0)  # 数量
        ws.cell(row=row_num, column=8, value=vals[6])  # 未税优惠单价
        ws.cell(row=row_num, column=9, value=vals[7])  # 含税单价
        ws.cell(row=row_num, column=10, value=vals[8])  # 含税优惠单价
        # 含税优惠合计 = 数量 * 含税优惠单价
        ws.cell(row=row_num, column=11, value=f'=G{row_num}*J{row_num}')
        for c in range(2, 12):
            ws.cell(row=row_num, column=c).border = thin_border

    # 备件费用
    ws['B25'] = '备件费用'
    ws['B25'].font = Font(bold=True)
    ws.merge_cells('B25:K25')
    ws['B25'].border = thin_border

    for row_num in range(26, 28):
        ws.cell(row=row_num, column=2, value=str(row_num - 25))
        ws.cell(row=row_num, column=11, value=f'=G{row_num}*J{row_num}')
        for c in range(2, 12):
            ws.cell(row=row_num, column=c).border = thin_border

    # 运输费用
    ws['B28'] = '运输费用'
    ws['B28'].font = Font(bold=True)
    ws.merge_cells('B28:K28')
    ws['B28'].border = thin_border

    ws.cell(row=29, column=2, value='1')
    ws.cell(row=29, column=11, value='=G29*J29')
    for c in range(2, 12):
        ws.cell(row=29, column=c).border = thin_border

    # 合计行
    ws['B30'] = '合计'
    ws.merge_cells('B30:J30')
    ws['B30'].font = Font(bold=True)
    ws['K30'] = '=SUM(K23:K24,K26:K27,K29)'  # 汇总所有明细行
    ws['B30'].border = thin_border
    ws['K30'].border = thin_border

    # 调整列宽
    col_widths = {'B': 12, 'C': 14, 'D': 12, 'E': 18, 'F': 12, 'G': 8, 'H': 16, 'I': 14, 'J': 16, 'K': 16}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 保存
    output_path = '服务报价单.xlsx'
    wb.save(output_path)
    print(f'✓ 已生成: {output_path}')
    return output_path


if __name__ == '__main__':
    create_quotation_template()
